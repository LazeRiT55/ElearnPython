import csv
import re
from collections import Counter
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl import Workbook
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader
import math

convert_currency = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class DataSet:
    def __init__(self, name, spec):
        self.file_name = name
        self.vacancies_objects, self.header = self.csv_reader()
        self.vacancies_objects = self.csv_filer()
        self.speciality = spec
        self.mean_salary = {}
        self.vacancies_count = {}
        self.spec_salary = {}
        self.spec_count = {}
        self.cities_salary = {}
        self.cities_salary_final = {}
        self.cities_vacancies_count = {}
        self.vacancy_fraction = {}


    def clear_string(text):
        text = re.sub(r'\<[^>]*\>', '', text)
        text = text.replace("\xa0", " ")
        text = text.replace(' ', " ")
        text = text.strip()
        text = re.sub(" +", " ", text)
        return text


    def csv_reader(self):
        file = open(self.file_name, encoding="utf-8-sig")
        reader = list(csv.reader(file))
        if not reader:
            print('Пустой файл')
            quit()
        header = reader[0]
        info = [x for x in reader[1:] if '' not in x and len(x) == len(header)]
        if len(info) == 0:
            print('Нет данных')
            quit()
        return info, header


    def csv_filer(self) -> list:
        output = []
        for i in range(len(self.vacancies_objects)):
            vacancy_info = []
            for j in range(len(self.vacancies_objects[i])):
                text = DataSet.clear_string(self.vacancies_objects[i][j])
                vacancy_info.append(text)
            output.append(vacancy_info)
        return output


    def find_year_item(self):
        vacancy_count = 0
        for vacancy in self.vacancies_objects:
            for i in range(len(vacancy)):
                if self.header[i] == 'name':
                    name = vacancy[i]
                if self.header[i] == 'salary_from':
                    salary_from = float(vacancy[i])
                elif self.header[i] == 'salary_to':
                    salary_to = float(vacancy[i])
                elif self.header[i] == 'salary_currency':
                    currency = vacancy[i]
                elif self.header[i] == 'area_name':
                    area = vacancy[i]
                    if area not in self.cities_salary.keys():
                        self.cities_salary[area] = []
                    if area not in self.cities_vacancies_count.keys():
                        self.cities_vacancies_count[area] = 0
                elif self.header[i] == 'published_at':
                    year = int(vacancy[i].split('-')[0])
                    if year not in self.mean_salary.keys():
                        self.mean_salary[year] = []
                    if year not in self.spec_salary.keys():
                        self.spec_salary[year] = []
                    if year not in self.vacancies_count.keys():
                        self.vacancies_count[year] = 0
                    if year not in self.spec_count.keys():
                        self.spec_count[year] = 0
            salary_average = int((salary_from + salary_to) // 2)
            salary_average = int(salary_average * convert_currency[currency])
            self.mean_salary[year].append(salary_average)
            self.cities_salary[area].append(salary_average)
            self.vacancies_count[year] += 1
            if self.speciality in name:
                self.spec_count[year] += 1
                self.spec_salary[year].append(salary_average)
            self.cities_vacancies_count[area] += 1
            vacancy_count += 1
        for year in self.mean_salary:
            if len(self.mean_salary[year]) == 0:
                self.mean_salary[year] = 0
            else:
                self.mean_salary[year] = int(sum(self.mean_salary[year]) // len(self.mean_salary[year]))
        for year in self.spec_salary:
            if len(self.spec_salary[year]) == 0:
                self.spec_salary[year] = 0
            else:
                self.spec_salary[year] = int(sum(self.spec_salary[year]) // len(self.spec_salary[year]))
        for area in self.cities_salary:
            if self.cities_vacancies_count[area] / vacancy_count >= 0.01 and len(self.cities_salary[area]) != 0:
                self.cities_salary_final[area] = int(sum(self.cities_salary[area]) // len(self.cities_salary[area]))
        for area in self.cities_vacancies_count:
            fraction = self.cities_vacancies_count[area] / vacancy_count
            if fraction >= 0.01:
                self.vacancy_fraction[area] = round(fraction, 4)


class Report:
    def __init__(self, profession):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Статистика по годам"
        self.ws1 = self.wb.create_sheet("Статистика по городам")
        self.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        self.fig, self.ax = plt.subplots(nrows=2, ncols=2, figsize=(10, 10))
        self.profession = profession
        self.env = Environment(loader=FileSystemLoader('.'))
        self.html = self.env.get_template("index.html")
        self.a = 1

    def set_column_width(self, name_column, sheet, count_rows, start):
        for i in range(len(name_column)):
            max_len = 0
            for j in range(1, count_rows + 1):
                if max_len < len(str(sheet[chr(i + ord(start)) + str(j)].value)):
                    max_len = len(str(sheet[chr(i + ord(start)) + str(j)].value)) + 2
            sheet.column_dimensions[chr(i + ord(start))].width = max_len

    def fill_header(self, boxes, name_column, is_bold):
        i = 0
        for column in boxes:
            for box in column:
                box.value = name_column[i]
                box.font = Font(bold=is_bold)
                box.border = self.border
                i += 1

    def generate_excel(self, list_name, data, header, start, end, type=""):
        self.ws = self.wb[list_name]
        self.fill_header(list(self.ws[start + '1': end + '1']), header, True)
        for i in range(len(data)):
            keys = list(data[i].keys())
            for j in range(len(keys)):
                self.ws[start + str(j + 2)].value = keys[j]
                self.ws[start + str(j + 2)].border = self.border
                self.ws[chr(i + ord(start) + 1) + str(j + 2)].value = data[i][keys[j]]
                self.ws[chr(i + ord(start) + 1) + str(j + 2)].border = self.border
                if type:
                    self.ws[chr(i + ord(start) + 1) + str(j + 2)].number_format = FORMAT_PERCENTAGE_00
        self.set_column_width(header, self.ws, len(list(data[0].keys())) + 1, start)

    def save_excel(self):
        self.wb.save("report.xlsx")

    def generate_image(self, salary_by_year, salary_by_year_for_profession, count_by_year, count_by_year_for_profession, sum_salary_by_year_for_city, fraction_by_city):
        width = 0.35
        x = np.array(list(salary_by_year.keys()))
        self.ax[0, 0].bar(x - width / 2, list(salary_by_year.values()), width, label='Средняя з/п')
        self.ax[0, 0].bar(x + width / 2, list(salary_by_year_for_profession.values()), width, label=f'з/п {spec}')
        self.ax[0, 0].set_title('Уровень зарплат по годам')
        self.ax[0, 0].set_xticks(x)
        self.ax[0, 0].set_xticklabels(list(salary_by_year.keys()), rotation=90, fontsize=8)
        self.ax[0, 0].legend(prop={'size': 8})
        self.ax[0, 0].grid(axis='y')
        self.ax[0, 1].bar(x - width / 2, list(count_by_year.values()), width, label='Количество вакансий')
        self.ax[0, 1].bar(x + width / 2, list(count_by_year_for_profession.values()), width, label=f'Количество вакансий {spec}')
        self.ax[0, 1].set_title('Количество вакансий по годам')
        self.ax[0, 1].set_xticks(x)
        self.ax[0, 1].set_xticklabels(list(salary_by_year.keys()), rotation=90, fontsize=8)
        self.ax[0, 1].legend(prop={'size': 8})
        self.ax[0, 1].grid(axis='y')
        label = []
        for item in list(sum_salary_by_year_for_city.keys()):
            if "-" in item:
                label.append(item.replace("-", "-\n"))
            elif " " in item:
                label.append(item.replace(" ", "\n"))
            else:
                label.append(item)
        y = np.array(label)
        self.ax[1, 0].barh(y, list(sum_salary_by_year_for_city.values()))
        self.ax[1, 0].set_yticks(y, labels=label, fontsize=6)
        self.ax[1, 0].invert_yaxis()
        self.ax[1, 0].set_title('Уровень зарплат по городам')
        self.ax[1, 0].grid(axis='x')
        summ = sum(list(fraction_by_city.values())[10:])
        fraction_by_city = dict(list(fraction_by_city.items())[0:10])
        fraction_by_city["Другие"] = summ
        self.ax[1, 1].pie(list(fraction_by_city.values()), labels=list(fraction_by_city.keys()), textprops={'fontsize': 6})
        self.ax[1, 1].axis("equal")
        self.ax[1, 1].set_title('Доля вакансий по городам')
        plt.tight_layout()
        plt.savefig('graph.png', dpi=100)

    def create_pdf(self):
        options = {
            "enable-local-file-access": None
        }
        xfile = openpyxl.load_workbook("report.xlsx")
        data = xfile['Статистика по годам']
        data2 = xfile['Статистика по городам']
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(self.html.render({'profession': self.profession, 'first': data, 'second': data2}), 'report.pdf', configuration=config, options=options)


name = input('Введите название файла: ')
spec = input('Введите название профессии: ')
data_set = DataSet(name, spec)
data_set.find_year_item()
xls = Report(spec)
header = ["Год", "Средняя зарплата", f"Средняя зарплата - {spec}", "Количество вакансий", f"Количество вакансий - {spec}"]
xls.generate_excel("Статистика по годам", [data_set.mean_salary, data_set.vacancies_count, data_set.spec_salary, data_set.spec_count], header, 'A', 'E')
header = ["Город", "Уровень зарплат"]
xls.generate_excel("Статистика по городам", [dict(Counter(data_set.cities_salary_final).most_common(10))], header, 'A', 'B')
header = ["Город", "Доля вакансий"]
xls.generate_excel("Статистика по городам", [dict(Counter(data_set.vacancy_fraction).most_common(10))], header, 'D', 'E', "percent")
xls.ws.column_dimensions["C"].width = 2
xls.save_excel()
xls.generate_image(data_set.mean_salary, data_set.spec_salary, data_set.vacancies_count, data_set.spec_count, dict(Counter(data_set.cities_salary_final).most_common(10)), dict(Counter(data_set.vacancy_fraction).most_common(10)))
xls.create_pdf()
